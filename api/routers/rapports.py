import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(__file__))))

import io
from datetime import date
from typing import Optional

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import and_, case, func, select
from sqlalchemy.orm import Session, joinedload

from api.deps import get_current_user, sbc_scope
from database import get_db
from models import CategorieEnum, Passage, SBCEnum, Site, StatutPassageEnum, User
from schemas.passage import PassageResponse

router = APIRouter()


def _filters(date_debut, date_fin, sbc, sto, statut, categorie, annee):
    f = []
    if date_debut:
        f.append(Passage.date_planifiee >= date_debut)
    if date_fin:
        f.append(Passage.date_planifiee <= date_fin)
    if annee:
        f.append(Passage.annee == annee)
    if sbc:
        f.append(Site.sbc == SBCEnum(sbc))
    if sto:
        f.append(Site.sto == sto)
    if statut:
        f.append(Passage.statut == statut)
    if categorie:
        f.append(Site.categorie == categorie)
    return f


@router.get("/passages", summary="Rapport des passages par plage de dates")
def rapport_passages(
    date_debut: Optional[date] = Query(None),
    date_fin: Optional[date] = Query(None),
    sbc: Optional[str] = Query(None),
    sto: Optional[str] = Query(None),
    statut: Optional[StatutPassageEnum] = Query(None),
    categorie: Optional[CategorieEnum] = Query(None),
    annee: Optional[int] = Query(None),
    skip: int = Query(0, ge=0),
    limit: int = Query(200, ge=1, le=1000),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    sbc = sbc_scope(sbc, current_user)
    today = date.today()
    flt = _filters(date_debut, date_fin, sbc, sto, statut, categorie, annee)

    base = select(Passage).join(Passage.site).where(*flt)
    total = db.execute(select(func.count()).select_from(base.subquery())).scalar()

    items = (
        db.execute(
            base.options(joinedload(Passage.site), joinedload(Passage.execution))
            .order_by(Passage.date_planifiee)
            .offset(skip)
            .limit(limit)
        )
        .scalars()
        .unique()
        .all()
    )

    row = db.execute(
        select(
            func.count().label("total"),
            func.sum(case((Passage.statut == StatutPassageEnum.Fait, 1), else_=0)).label("faits"),
            func.sum(case((Passage.statut == StatutPassageEnum.Non_effectue, 1), else_=0)).label("non_effectues"),
            func.sum(
                case((and_(Passage.statut == StatutPassageEnum.Non_effectue, Passage.date_planifiee < today), 1), else_=0)
            ).label("en_retard"),
        )
        .select_from(Passage)
        .join(Passage.site)
        .where(*flt)
    ).one()

    t, f, ne, er = row.total or 0, row.faits or 0, row.non_effectues or 0, row.en_retard or 0

    return {
        "total": total,
        "stats": {
            "total": t,
            "faits": f,
            "non_effectues": ne,
            "en_retard": er,
            "a_venir": t - f - ne,
            "taux_realisation": round(f / t * 100, 1) if t > 0 else 0.0,
        },
        "skip": skip,
        "limit": limit,
        "items": [PassageResponse.model_validate(p) for p in items],
    }


@router.get("/passages/export-excel", summary="Export Excel du rapport de passages")
def export_passages_excel(
    date_debut: Optional[date] = Query(None),
    date_fin: Optional[date] = Query(None),
    sbc: Optional[str] = Query(None),
    sto: Optional[str] = Query(None),
    statut: Optional[StatutPassageEnum] = Query(None),
    categorie: Optional[CategorieEnum] = Query(None),
    annee: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    sbc = sbc_scope(sbc, current_user)
    today = date.today()
    flt = _filters(date_debut, date_fin, sbc, sto, statut, categorie, annee)

    passages = (
        db.execute(
            select(Passage)
            .join(Passage.site)
            .options(joinedload(Passage.site), joinedload(Passage.execution))
            .where(*flt)
            .order_by(Passage.date_planifiee)
        )
        .scalars()
        .unique()
        .all()
    )

    total = len(passages)
    faits = sum(1 for p in passages if p.statut == StatutPassageEnum.Fait)
    non_eff = sum(1 for p in passages if p.statut == StatutPassageEnum.Non_effectue)
    en_retard = sum(1 for p in passages if p.statut == StatutPassageEnum.Non_effectue and p.date_planifiee < today)
    a_venir = total - faits - non_eff
    taux = round(faits / total * 100, 1) if total > 0 else 0.0

    wb = openpyxl.Workbook()

    # ── Feuille 1 : Synthèse ──────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Synthèse"

    hdr_fill = PatternFill("solid", fgColor="1F3864")
    hdr_font = Font(color="FFFFFF", bold=True, size=11)
    center = Alignment(horizontal="center", vertical="center")

    ws1.merge_cells("A1:B1")
    ws1["A1"] = "Rapport de Maintenance Préventive — PM MTN CI"
    ws1["A1"].font = Font(color="1F3864", bold=True, size=13)
    ws1["A1"].alignment = center
    ws1.row_dimensions[1].height = 22

    ws1["A3"] = "Période"
    ws1["B3"] = f"{date_debut or '—'} → {date_fin or '—'}"
    ws1["A4"] = "Généré le"
    ws1["B4"] = today.strftime("%d/%m/%Y")
    if sbc:
        ws1["A5"] = "SBC"
        ws1["B5"] = sbc

    ws1["A7"] = "Indicateur"
    ws1["B7"] = "Valeur"
    for col in ["A7", "B7"]:
        ws1[col].fill = hdr_fill
        ws1[col].font = hdr_font
        ws1[col].alignment = center

    stats_rows = [
        ("Total passages", total),
        ("Effectués (Faits)", faits),
        ("Non effectués", non_eff),
        ("En retard", en_retard),
        ("À venir", a_venir),
        ("Taux de réalisation", f"{taux} %"),
    ]
    alt_fill = PatternFill("solid", fgColor="EEF2FF")
    for i, (label, val) in enumerate(stats_rows, start=8):
        ws1[f"A{i}"] = label
        ws1[f"B{i}"] = val
        if i % 2 == 0:
            ws1[f"A{i}"].fill = alt_fill
            ws1[f"B{i}"].fill = alt_fill

    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions["B"].width = 20

    # ── Feuille 2 : Détail ────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Détail")

    headers = [
        "Code site", "Nom", "Catégorie", "SBC", "STO",
        "Passage", "Mois", "Année", "Date planifiée",
        "Statut", "Date exécution", "WO / Ticket",
    ]
    col_widths = [12, 28, 12, 10, 24, 10, 12, 8, 16, 14, 16, 16]

    for ci, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws2.cell(row=1, column=ci, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = center
        ws2.column_dimensions[get_column_letter(ci)].width = w
    ws2.row_dimensions[1].height = 18

    stripe_fill = PatternFill("solid", fgColor="F8FAFC")
    for ri, p in enumerate(passages, start=2):
        values = [
            p.site.code_site,
            p.site.nom,
            p.site.categorie.value,
            p.site.sbc.value,
            p.site.sto,
            f"{p.passage_num}/{p.total_passages}",
            p.mois_nom,
            p.annee,
            p.date_planifiee.strftime("%d/%m/%Y"),
            p.statut.value,
            p.execution.date_execution.strftime("%d/%m/%Y") if p.execution else "",
            p.execution.wo_ticket if p.execution else "",
        ]
        for ci, val in enumerate(values, start=1):
            cell = ws2.cell(row=ri, column=ci, value=val)
            if ri % 2 == 0:
                cell.fill = stripe_fill

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    d_start = date_debut.isoformat() if date_debut else "debut"
    d_end = date_fin.isoformat() if date_fin else "fin"
    filename = f"Rapport_PM_MTN_{d_start}_{d_end}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
